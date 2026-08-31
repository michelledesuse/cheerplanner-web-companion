"""Backend tests for iter107: Scouting Skill Library bulk upload + selection model.

Covers:
- GET /api/team/scouting/skills/template?fmt=csv|xlsx
- POST /api/team/scouting/skills/import (.csv & .xlsx, aliases, clamp, dedupe, invalid)
- GET /api/team/scouting/report/{roster_id} include_all coach vs filtered parent/athlete
- PUT ...skill/{skill_id} setting + clearing level
"""
import io
import os
import pytest
import requests
from openpyxl import Workbook, load_workbook

# Load EXPO_PUBLIC_BACKEND_URL from frontend/.env when running pytest from CLI
if not os.environ.get("EXPO_PUBLIC_BACKEND_URL"):
    try:
        with open("/app/frontend/.env") as _f:
            for _l in _f:
                if _l.startswith("EXPO_PUBLIC_BACKEND_URL="):
                    os.environ["EXPO_PUBLIC_BACKEND_URL"] = _l.split("=", 1)[1].strip().strip('"')
                    break
    except Exception:
        pass

BASE_URL = os.environ.get("EXPO_PUBLIC_BACKEND_URL", "").rstrip("/")
assert BASE_URL, "EXPO_PUBLIC_BACKEND_URL missing"

COACH = ("coach.casey@cheerplanner.app", "CheerDemo2026!")
DEMO = ("demo@cheerplanner.app", "CheerDemo2026!")
PARENT = ("parent.taylor@cheerplanner.app", "CheerDemo2026!")
ATHLETE = ("sophia.athlete@cheerplanner.app", "CheerDemo2026!")


def _login(session: requests.Session, email: str, password: str) -> str:
    r = session.post(f"{BASE_URL}/api/auth/login", json={"email": email, "password": password})
    assert r.status_code == 200, f"login failed for {email}: {r.status_code} {r.text}"
    tok = r.json().get("token") or r.json().get("access_token")
    assert tok, f"no token returned: {r.json()}"
    return tok


@pytest.fixture(scope="module")
def coach_headers():
    s = requests.Session()
    tok = _login(s, *COACH)
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="module")
def demo_headers():
    s = requests.Session()
    tok = _login(s, *DEMO)
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="module")
def parent_headers():
    s = requests.Session()
    tok = _login(s, *PARENT)
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="module")
def athlete_headers():
    s = requests.Session()
    tok = _login(s, *ATHLETE)
    return {"Authorization": f"Bearer {tok}"}


# ------------------------- Template downloads -------------------------
class TestTemplate:
    def test_template_csv(self, coach_headers):
        r = requests.get(f"{BASE_URL}/api/team/scouting/skills/template?fmt=csv", headers=coach_headers)
        assert r.status_code == 200, r.text
        assert "text/csv" in r.headers.get("content-type", "")
        assert "attachment" in r.headers.get("content-disposition", "").lower()
        body = r.text
        first_line = body.splitlines()[0]
        assert first_line == "Category,Level,Skill Name", f"Header mismatch: {first_line!r}"

    def test_template_xlsx(self, coach_headers):
        r = requests.get(f"{BASE_URL}/api/team/scouting/skills/template?fmt=xlsx", headers=coach_headers)
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert headers == ["Category", "Level", "Skill Name"], f"XLSX header mismatch: {headers}"

    def test_template_requires_team_access(self):
        # Parent (no team_access) should get 403
        s = requests.Session()
        tok = _login(s, *PARENT)
        r = requests.get(
            f"{BASE_URL}/api/team/scouting/skills/template?fmt=csv",
            headers={"Authorization": f"Bearer {tok}"},
        )
        assert r.status_code in (401, 403), f"expected 403 for parent, got {r.status_code}"


# ------------------------- Import CSV / XLSX -------------------------
def _existing_skill(coach_headers):
    """Return an existing (category, level, name) already in the library — for dedupe."""
    r = requests.get(f"{BASE_URL}/api/team/scouting/skills", headers=coach_headers)
    assert r.status_code == 200, r.text
    cats = r.json().get("categories") or {}
    for cat, arr in cats.items():
        if arr:
            s = arr[0]
            return cat, int(s.get("level_group") or 1), s.get("name")
    pytest.skip("No existing skills in library to test dedupe")


class TestImport:
    def test_import_csv_with_aliases_clamp_and_dedupe(self, coach_headers):
        cat, lvl, dup_name = _existing_skill(coach_headers)
        # Build small CSV: 2 new + 1 duplicate + 1 alias/clamp + 1 invalid
        csv_body = (
            "Category,Level,Skill Name\n"
            "Tumble,2,TEST_iter107_ScoutSkill_A\n"  # alias 'Tumble' -> tumbling
            "Stunt,9,TEST_iter107_ScoutSkill_B\n"   # clamp 9 -> 7, alias 'Stunt' -> stunting
            f"{cat},{lvl},{dup_name}\n"              # duplicate
            ",1,Missing category\n"                  # invalid
            "Jumps,,MissingLevel\n"                  # invalid
        )
        files = {"file": ("scouting_import_test.csv", csv_body, "text/csv")}
        r = requests.post(
            f"{BASE_URL}/api/team/scouting/skills/import",
            headers=coach_headers, files=files,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("added") == 2, f"expected 2 added, got {data}"
        assert data.get("skipped_duplicates") == 1, f"expected 1 duplicate, got {data}"
        assert data.get("invalid_rows") == 2, f"expected 2 invalid, got {data}"

        # Verify skills appear under correct category+level
        r2 = requests.get(f"{BASE_URL}/api/team/scouting/skills", headers=coach_headers)
        assert r2.status_code == 200
        cats = r2.json()["categories"]
        tumbling_names = [(s["name"], s.get("level_group")) for s in cats.get("tumbling", [])]
        stunting_names = [(s["name"], s.get("level_group")) for s in cats.get("stunting", [])]
        assert ("TEST_iter107_ScoutSkill_A", 2) in tumbling_names, \
            f"Missing new skill A in tumbling@2: {tumbling_names[:5]}..."
        assert ("TEST_iter107_ScoutSkill_B", 7) in stunting_names, \
            f"Missing new skill B in stunting@7 (clamp): {stunting_names[:5]}..."

    def test_import_xlsx_works(self, coach_headers):
        wb = Workbook()
        ws = wb.active
        ws.append(["Category", "Level", "Skill Name"])
        ws.append(["Jumps", 3, "TEST_iter107_XlsxJump_C"])
        buf = io.BytesIO()
        wb.save(buf)
        files = {"file": ("scouting_import.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(
            f"{BASE_URL}/api/team/scouting/skills/import",
            headers=coach_headers, files=files,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("added") >= 1, data

        r2 = requests.get(f"{BASE_URL}/api/team/scouting/skills", headers=coach_headers)
        jumps_names = [(s["name"], s.get("level_group")) for s in r2.json()["categories"].get("jumps", [])]
        assert ("TEST_iter107_XlsxJump_C", 3) in jumps_names, f"Missing XLSX skill: {jumps_names[:5]}..."


# ------------------------- Selection model -------------------------
def _find_athlete_roster(headers, prefer_first_name=None):
    r = requests.get(f"{BASE_URL}/api/team/scouting/overview", headers=headers)
    assert r.status_code == 200, r.text
    ov = r.json()
    ath = ov.get("athletes") or []
    if prefer_first_name:
        for a in ath:
            if (a.get("first_name") or "").lower() == prefer_first_name.lower():
                return a["roster_id"]
    return ath[0]["roster_id"] if ath else None


class TestSelectionModel:
    def test_coach_sees_all_skills_include_all(self, coach_headers):
        roster_id = _find_athlete_roster(coach_headers)
        assert roster_id, "no athlete in coach overview"
        r = requests.get(f"{BASE_URL}/api/team/scouting/report/{roster_id}", headers=coach_headers)
        assert r.status_code == 200, r.text
        rep = r.json()
        assert rep["role"] == "coach"
        assert rep["can_edit"] is True
        total = sum(len(v) for v in rep["categories"].values())
        # Demo hub seed = 216 skills across categories/levels
        assert total >= 200, f"coach view returned only {total} skills (expected all library)"

    def test_parent_and_athlete_see_only_selected(self, coach_headers, parent_headers, athlete_headers):
        roster_id = _find_athlete_roster(parent_headers) or _find_athlete_roster(athlete_headers, "sophia")
        assert roster_id, "parent/athlete cannot see any roster"

        # Get a skill to select
        r = requests.get(f"{BASE_URL}/api/team/scouting/skills", headers=coach_headers)
        cats = r.json()["categories"]
        skill = None
        for cat, arr in cats.items():
            if arr:
                skill = arr[0]
                break
        assert skill, "no skill available"
        skill_id = skill["id"]

        # Baseline parent view
        rp0 = requests.get(f"{BASE_URL}/api/team/scouting/report/{roster_id}", headers=parent_headers)
        assert rp0.status_code == 200, rp0.text
        base_total = sum(len(v) for v in rp0.json()["categories"].values())

        # Coach sets a level -> "select"
        rp = requests.put(
            f"{BASE_URL}/api/team/scouting/report/{roster_id}/skill/{skill_id}",
            headers=coach_headers, json={"level": "spotted", "notes": "iter107 test note"},
        )
        assert rp.status_code == 200, rp.text

        # Parent view: only selected skills; must include this skill, every returned has a level
        rp1 = requests.get(f"{BASE_URL}/api/team/scouting/report/{roster_id}", headers=parent_headers)
        assert rp1.status_code == 200, rp1.text
        rep1 = rp1.json()
        assert rep1["role"] == "parent", f"expected parent, got {rep1.get('role')}"
        assert rep1["can_edit"] is False
        all_skills = [s for arr in rep1["categories"].values() for s in arr]
        assert any(s["skill_id"] == skill_id and s["level"] == "spotted" for s in all_skills), \
            "selected skill missing from parent view"
        assert all(s.get("level") for s in all_skills), \
            "parent view should ONLY include skills with a level set"

        # Athlete view — try Sophia (has ATHLETE chat link approved)
        soph_roster = _find_athlete_roster(athlete_headers, "sophia")
        if soph_roster:
            # If the coach-selected roster matches Sophia, use it; else set on Sophia too
            target = soph_roster if soph_roster == roster_id else None
            if target is None:
                # Also select on Sophia's roster
                requests.put(
                    f"{BASE_URL}/api/team/scouting/report/{soph_roster}/skill/{skill_id}",
                    headers=coach_headers, json={"level": "spotted"},
                )
                target = soph_roster
            ra = requests.get(f"{BASE_URL}/api/team/scouting/report/{target}", headers=athlete_headers)
            assert ra.status_code == 200, ra.text
            rep_a = ra.json()
            assert rep_a["role"] == "athlete"
            ath_skills = [s for arr in rep_a["categories"].values() for s in arr]
            assert all(s.get("level") for s in ath_skills), "athlete view should be filtered too"
            assert any(s["skill_id"] == skill_id for s in ath_skills), "selected skill missing in athlete view"

        # Clearing the level removes it from parent/athlete view
        rc = requests.put(
            f"{BASE_URL}/api/team/scouting/report/{roster_id}/skill/{skill_id}",
            headers=coach_headers, json={"level": ""},
        )
        assert rc.status_code == 200, rc.text

        rp2 = requests.get(f"{BASE_URL}/api/team/scouting/report/{roster_id}", headers=parent_headers)
        rep2 = rp2.json()
        all_skills2 = [s for arr in rep2["categories"].values() for s in arr]
        assert not any(s["skill_id"] == skill_id for s in all_skills2), \
            "cleared skill should NOT appear in parent view after removal"
        assert len(all_skills2) <= max(base_total, 0), "parent view leaks unselected skills after clear"
