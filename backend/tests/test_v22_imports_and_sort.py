"""Backend tests for CheerPlanner v2.2 imports (teams_to_watch, XLSX templates) and sort regression."""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("EXPO_PUBLIC_BACKEND_URL") or os.environ.get("EXPO_BACKEND_URL")
assert BASE_URL, "EXPO_PUBLIC_BACKEND_URL must be set"
BASE_URL = BASE_URL.rstrip("/")
API = f"{BASE_URL}/api"

EMAIL = "applereview@cheerplanner.app"
PASSWORD = "Review2026!"


@pytest.fixture(scope="module")
def token():
    r = requests.post(f"{API}/auth/login", json={"email": EMAIL, "password": PASSWORD}, timeout=30)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    body = r.json()
    return body.get("access_token") or body.get("token")


@pytest.fixture(scope="module")
def headers(token):
    return {"Authorization": f"Bearer {token}"}


# ---------- CSV / XLSX templates ----------
def test_csv_template_teams_to_watch(headers):
    r = requests.get(f"{API}/import/template/teams_to_watch?fmt=csv", headers=headers, timeout=30)
    assert r.status_code == 200
    assert "text/csv" in r.headers.get("content-type", "")
    first_line = r.text.splitlines()[0]
    for col in ["Competition", "Team Name", "Date", "Location", "Performance Time"]:
        assert col in first_line, f"missing {col} in CSV header"


@pytest.mark.parametrize("kind", ["teams_to_watch", "expenses", "competitions", "travel", "schedule"])
def test_xlsx_templates_valid(headers, kind):
    r = requests.get(f"{API}/import/template/{kind}?fmt=xlsx", headers=headers, timeout=30)
    assert r.status_code == 200
    assert r.headers.get("content-type") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    wb = load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    assert "Template" in wb.sheetnames


def test_xlsx_expenses_has_reference_sheet_with_14_categories(headers):
    r = requests.get(f"{API}/import/template/expenses?fmt=xlsx", headers=headers, timeout=30)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    assert "Reference" in wb.sheetnames, f"Reference sheet missing. Got {wb.sheetnames}"
    ref = wb["Reference"]
    lines = []
    for row in ref.iter_rows(values_only=True):
        for c in row:
            if c is not None and str(c).strip():
                lines.append(str(c).strip())
    # Categories appear as bullet lines starting with •
    cats = [ln.lstrip("•").strip() for ln in lines if ln.startswith("•")]
    assert len(cats) == 14, f"expected 14 categories, got {len(cats)}: {cats}"


def test_xlsx_teams_to_watch_template_headers(headers):
    r = requests.get(f"{API}/import/template/teams_to_watch?fmt=xlsx", headers=headers, timeout=30)
    wb = load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    ws = wb["Template"]
    headers_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for col in ["Competition", "Team Name", "Date", "Location", "Performance Time"]:
        assert col in headers_row


def test_unknown_template_returns_400(headers):
    r = requests.get(f"{API}/import/template/bogus?fmt=csv", headers=headers, timeout=30)
    assert r.status_code == 400


# ---------- Preview ----------
def test_preview_teams_to_watch_csv(headers):
    csv_data = (
        "Competition,Team Name,Date,Location,Performance Time\n"
        "TEST_v22 Watch Comp,TEST_Panthers,2025-11-14,Floor A,2:30 PM\n"
        "TEST_v22 Watch Comp,TEST_Top Gun,2025-11-15,Floor B,11:00 AM\n"
        ",NoCompTeam,2025-11-15,Floor C,11:00 AM\n"
    )
    files = {"file": ("teams.csv", csv_data, "text/csv")}
    data = {"kind": "teams_to_watch"}
    r = requests.post(f"{API}/import/preview", headers=headers, files=files, data=data, timeout=30)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["kind"] == "teams_to_watch"
    assert "existing_competitions" in body
    assert isinstance(body["existing_competitions"], list)
    # parser keeps rows even with empty competition; the commit step will skip them
    assert len(body["rows"]) >= 2
    names = [row["name"] for row in body["rows"]]
    assert "TEST_Panthers" in names


# ---------- Commit teams_to_watch ----------
@pytest.fixture(scope="module")
def created_comp_id(headers):
    """Create a known competition for matched-row testing; clean up at end."""
    payload = {
        "name": "TEST_v22 Existing Comp",
        "event_date": "2025-12-01",
        "location": "TestCity",
        "housing_required": False,
    }
    r = requests.post(f"{API}/competitions", headers=headers, json=payload, timeout=30)
    assert r.status_code in (200, 201), r.text
    cid = r.json()["id"]
    yield cid
    requests.delete(f"{API}/competitions/{cid}", headers=headers, timeout=30)


def test_commit_teams_to_watch_matches_and_auto_creates(headers, created_comp_id):
    rows = [
        # matched by name (case insensitive)
        {"competition": "test_v22 existing comp", "name": "TEST_MatchedTeam", "date": "2025-12-01", "location": "Floor X", "performance_time": "14:30"},
        # unmatched -> auto-create placeholder
        {"competition": "TEST_v22 Placeholder Comp", "name": "TEST_PlaceholderTeam", "date": None, "location": None, "performance_time": None},
        # missing competition -> skipped
        {"competition": "", "name": "TEST_Skip", "date": None, "location": None, "performance_time": None},
        # missing name -> skipped
        {"competition": "TEST_v22 Existing Comp", "name": "", "date": None, "location": None, "performance_time": None},
    ]
    r = requests.post(f"{API}/import/commit", headers=headers,
                      json={"kind": "teams_to_watch", "rows": rows}, timeout=30)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 2, body
    assert body["skipped"] == 2, body
    assert any("Placeholder" in w or "placeholder" in w for w in body["warnings"]), body["warnings"]

    # Verify matched team appended to existing competition
    rc = requests.get(f"{API}/competitions/{created_comp_id}", headers=headers, timeout=30)
    assert rc.status_code == 200
    comp = rc.json()
    tw_names = [t["name"] for t in comp.get("teams_to_watch", [])]
    assert "TEST_MatchedTeam" in tw_names

    # Verify placeholder competition was created and contains TEST_PlaceholderTeam
    rl = requests.get(f"{API}/competitions", headers=headers, timeout=30)
    assert rl.status_code == 200
    placeholder = None
    for c in rl.json():
        if c["name"] == "TEST_v22 Placeholder Comp":
            placeholder = c
            break
    assert placeholder is not None, "placeholder competition not created"
    tw_names_p = [t["name"] for t in placeholder.get("teams_to_watch", [])]
    assert "TEST_PlaceholderTeam" in tw_names_p

    # Cleanup placeholder
    requests.delete(f"{API}/competitions/{placeholder['id']}", headers=headers, timeout=30)


# ---------- Regression: existing import kinds still work ----------
def test_preview_competitions_csv(headers):
    csv_data = (
        "Competition Name,Location,Event Date\n"
        "TEST_v22 Regression Comp,TestVille,2026-02-01\n"
    )
    files = {"file": ("comps.csv", csv_data, "text/csv")}
    data = {"kind": "competitions"}
    r = requests.post(f"{API}/import/preview", headers=headers, files=files, data=data, timeout=30)
    assert r.status_code == 200, r.text
    assert r.json()["count"] == 1


def test_preview_expenses_long_csv(headers):
    csv_data = (
        "Date,Athlete,Category,Amount,Due Date,Paid\n"
        "2025-10-01,TEST_v22_Athlete,Tuition,250.00,2025-10-05,No\n"
    )
    files = {"file": ("exp.csv", csv_data, "text/csv")}
    data = {"kind": "expenses"}
    r = requests.post(f"{API}/import/preview", headers=headers, files=files, data=data, timeout=30)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["format"] == "long"
    assert body["count"] == 1


def test_preview_schedule_csv(headers):
    csv_data = (
        "Title,Type,Date,Start Time,End Time,Location,Athletes\n"
        "TEST_v22 Practice,Practice,2025-11-01,6:30 PM,8:30 PM,Gym,Ava\n"
    )
    files = {"file": ("sched.csv", csv_data, "text/csv")}
    data = {"kind": "schedule"}
    r = requests.post(f"{API}/import/preview", headers=headers, files=files, data=data, timeout=30)
    assert r.status_code == 200, r.text
    assert r.json()["count"] == 1


def test_preview_travel_csv(headers):
    csv_data = (
        "Competition,Hotel Name,Check In,Check Out,Hotel Cost\n"
        "TEST_v22 Existing Comp,Hyatt,2025-12-01,2025-12-03,500\n"
    )
    files = {"file": ("travel.csv", csv_data, "text/csv")}
    data = {"kind": "travel"}
    r = requests.post(f"{API}/import/preview", headers=headers, files=files, data=data, timeout=30)
    assert r.status_code == 200, r.text
    assert r.json()["count"] == 1


def test_preview_unknown_kind_400(headers):
    files = {"file": ("x.csv", "a,b\n1,2\n", "text/csv")}
    r = requests.post(f"{API}/import/preview", headers=headers, files=files, data={"kind": "bogus"}, timeout=30)
    assert r.status_code == 400
