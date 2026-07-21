"""Spreadsheet import helpers — parses CSV / XLSX into normalized row dicts."""
import csv
import io
import re
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Header normalization
# ---------------------------------------------------------------------------
def _norm(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return s.strip()


# Map of normalized synonyms → canonical key
COMPETITION_HEADERS = {
    "competition": "name", "competition name": "name", "name": "name", "event": "name", "event name": "name", "comp": "name",
    "location": "location", "city": "location", "venue": "location",
    "event date": "event_date", "date": "event_date", "competition date": "event_date", "start date": "event_date",
    "travel dates": "travel_dates",
    "end date": "end_date", "ends": "end_date",
    "housing required": "housing_required", "housing required ": "housing_required", "housing": "housing_required",
    "stay to play": "housing_required",
    "booking link": "booking_link", "link": "booking_link", "hotel link": "booking_link",
    "gym link": "booking_link", "general link": "booking_link",
    "booking release date and time": "booking_release_at",
    "booking release": "booking_release_at", "release date": "booking_release_at",
    "link release date": "booking_release_at", "link release": "booking_release_at",
    "link release time": "booking_release_time",
    "booking opens": "booking_release_at", "rooms open": "booking_release_at",
    "notes": "notes", "note": "notes",
}

TRAVEL_HEADERS = {
    "competition": "competition", "competition name": "competition", "comp": "competition", "event": "competition", "event name": "competition",
    "dates": "competition_dates", "travel day": "travel_day",
    "team": "team", "team s": "team",
    "hotel name": "hotel_provider", "hotel": "hotel_provider",
    "hotel confirmation": "hotel_confirmation", "hotel conf": "hotel_confirmation",
    "hotel confirmation number": "hotel_confirmation", "conf": "hotel_confirmation",
    "check in": "check_in", "check in date": "check_in", "checkin": "check_in",
    "check out": "check_out", "check out date": "check_out", "checkout": "check_out",
    "cancel date": "cancel_by", "cancel by": "cancel_by", "free cancel by": "cancel_by",
    "hotel cost": "hotel_cost", "hotel total": "hotel_cost", "hotel price": "hotel_cost",
    "cost": "hotel_cost", "total cost": "hotel_cost",
    "hotel paid": "hotel_paid", "hotel amount paid": "hotel_paid", "amount paid": "hotel_paid", "paid": "hotel_paid",
    "hotel balance due date": "hotel_due", "hotel balance due": "hotel_due", "balance due date": "hotel_due", "balance due": "hotel_due",
    # rental car
    "rental car company": "car_provider", "rental car": "car_provider", "car company": "car_provider", "car": "car_provider",
    "rental car confirmation": "car_confirmation", "car confirmation": "car_confirmation", "rental conf": "car_confirmation",
    "rental car cost": "car_cost", "car cost": "car_cost",
    "pickup date": "car_pickup_date", "pick up date": "car_pickup_date", "pick-up date": "car_pickup_date",
    "pickup time": "car_pickup_time", "pick up time": "car_pickup_time", "pick-up time": "car_pickup_time",
    "pickup location": "car_pickup_location", "pick up location": "car_pickup_location", "pick-up location": "car_pickup_location",
    "dropoff date": "car_dropoff_date", "drop off date": "car_dropoff_date", "drop-off date": "car_dropoff_date",
    "dropoff time": "car_dropoff_time", "drop off time": "car_dropoff_time", "drop-off time": "car_dropoff_time",
    "dropoff location": "car_dropoff_location", "drop off location": "car_dropoff_location", "drop-off location": "car_dropoff_location",
    # flights (outbound + return separated)
    "airline": "airline", "outbound airline": "airline",
    "flight confirmation": "flight_confirmation", "flight conf": "flight_confirmation", "outbound confirmation": "flight_confirmation",
    "flight number": "flight_number", "outbound flight number": "flight_number",
    "depart time": "depart_time", "departure time": "depart_time", "departure": "depart_time",
    "arrive time": "arrive_time", "arrival time": "arrive_time", "arrival": "arrive_time",
    "outbound cost": "outbound_cost",
    "flight cost": "flight_cost", "airfare": "flight_cost", "total flight cost": "flight_cost",
    "flight paid": "flight_paid", "flight amount paid": "flight_paid",
    "return airline": "return_airline",
    "return confirmation": "return_confirmation", "return flight confirmation": "return_confirmation",
    "return flight number": "return_flight_number", "return flight": "return_flight_number",
    "return depart time": "return_depart_time", "return departure": "return_depart_time",
    "return arrive time": "return_arrive_time", "return arrival": "return_arrive_time",
    "return cost": "return_cost",
}

EXPENSE_HEADERS = {
    "date": "date", "incurred on": "date", "month": "date",
    "athlete": "athlete", "athlete name": "athlete", "kid": "athlete", "child": "athlete",
    "category": "category", "type": "category", "expense": "category", "expense type": "category",
    "amount": "amount", "cost": "amount", "price": "amount",
    "due date": "due_date", "due": "due_date",
    "paid": "paid", "is paid": "paid",
    "note": "note", "notes": "note", "memo": "note", "description": "note",
}

SCHEDULE_HEADERS = {
    "title": "title", "event": "title", "event name": "title", "name": "title",
    "type": "event_type", "event type": "event_type", "kind": "event_type", "category": "event_type",
    "date": "date", "event date": "date", "start date": "date", "starts": "date",
    "start time": "start_time", "start": "start_time", "time": "start_time", "begin": "start_time",
    "end time": "end_time", "end": "end_time", "finish": "end_time",
    "location": "location", "where": "location", "gym": "location", "venue": "location",
    "athletes": "athletes", "athlete": "athletes", "kids": "athletes", "kid": "athletes",
    "children": "athletes", "child": "athletes", "for": "athletes",
    "repeats": "repeats", "repeat": "repeats", "frequency": "repeats", "recurrence": "repeats",
    "repeat days": "repeat_days", "days of week": "repeat_days", "days": "repeat_days",
    "on days": "repeat_days", "weekdays": "repeat_days",
    "repeat until": "repeat_until", "until": "repeat_until", "ends": "repeat_until",
    "repeats until": "repeat_until", "end date": "repeat_until",
    "notes": "notes", "note": "notes", "memo": "notes", "description": "notes",
}

KNOWN_CATEGORIES = [
    "Tuition", "Practice", "Gear", "Comp/Choreo", "Camp", "Uniform",
    "Classes & Privates", "Bow", "Warm-Up & Bag", "End of Season Comp Fees",
    "Registration", "Membership", "Late Fees", "Misc",
]

TEAMS_TO_WATCH_HEADERS = {
    "competition": "competition", "competition name": "competition", "comp": "competition",
    "event": "competition", "event name": "competition",
    "team": "name", "team name": "name", "name": "name", "watch": "name", "team to watch": "name",
    "date": "date", "performance date": "date", "watch date": "date",
    "location": "location", "venue": "location", "floor": "location", "stage": "location",
    "time": "performance_time", "performance time": "performance_time",
    "perform time": "performance_time", "watch time": "performance_time",
}


def _bool_from(v: Any) -> Optional[bool]:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("yes", "y", "true", "1", "x", "✓", "paid"):
        return True
    if s in ("no", "n", "false", "0", "", "unpaid"):
        return False
    return None


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _date(v: Any) -> Optional[str]:
    """Return YYYY-MM-DD if parseable, else None."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d",
        "%m-%d-%Y", "%d %b %Y", "%b %d %Y", "%B %d, %Y",
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M",
    ):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # last try: ISO parse
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date().isoformat()
    except Exception:
        return None


def _datetime(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
        "%m/%d/%Y %I:%M %p", "%m/%d/%Y %H:%M",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(s, fmt).isoformat()
        except ValueError:
            continue
    return None


def _combine_dt(date_val: Any, time_val: Any) -> Optional[str]:
    """Build a `YYYY-MM-DD HH:mm` string from a date cell + a time cell.

    Either may be missing. Returns just the date or time if the other is empty.
    """
    iso_date = _date(date_val) if date_val else None
    hhmm = _time24(time_val) if time_val else None
    if iso_date and hhmm:
        return f"{iso_date} {hhmm}"
    return iso_date or hhmm or None


# ---------------------------------------------------------------------------
# File reading
# ---------------------------------------------------------------------------
def read_table(filename: str, content: bytes) -> List[List[List[Any]]]:
    """Return list of sheets, each a list of rows, each a list of cells.

    For CSV, returns a single 'sheet'.
    For XLSX, returns one entry per non-empty worksheet.
    """
    name_lower = filename.lower()
    sheets: List[List[List[Any]]] = []
    if name_lower.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        sheets.append([list(row) for row in reader])
        return sheets
    if name_lower.endswith(".xlsx") or name_lower.endswith(".xlsm"):
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        for ws in wb.worksheets:
            rows = []
            for r in ws.iter_rows(values_only=True):
                # skip fully-empty rows
                if not any(c not in (None, "") for c in r):
                    continue
                rows.append(list(r))
            if rows:
                sheets.append(rows)
        return sheets
    raise ValueError("Unsupported file type. Please upload .csv or .xlsx.")


def _find_header_row(rows: List[List[Any]], header_map: Dict[str, str]) -> int:
    """Find row index that looks most like headers (max # of recognized headers)."""
    best = (0, 0)
    for i, row in enumerate(rows[:10]):
        hits = sum(1 for c in row if _norm(c) in header_map)
        if hits > best[0]:
            best = (hits, i)
    return best[1]


def _row_to_dict(headers: List[str], values: List[Any], header_map: Dict[str, str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for h, v in zip(headers, values):
        key = header_map.get(_norm(h))
        if key and out.get(key) in (None, ""):
            out[key] = v
    return out


# ---------------------------------------------------------------------------
# Competitions
# ---------------------------------------------------------------------------
def parse_competitions(filename: str, content: bytes) -> List[Dict[str, Any]]:
    sheets = read_table(filename, content)
    out: List[Dict[str, Any]] = []
    for rows in sheets:
        if not rows:
            continue
        hdr_idx = _find_header_row(rows, COMPETITION_HEADERS)
        headers = [str(c) if c is not None else "" for c in rows[hdr_idx]]
        for r in rows[hdr_idx + 1:]:
            rec = _row_to_dict(headers, r, COMPETITION_HEADERS)
            if not rec.get("name"):
                continue
            housing = rec.get("housing_required")
            housing_bool = _bool_from(housing) if housing is not None else False
            out.append({
                "name": str(rec.get("name")).strip(),
                "location": str(rec["location"]).strip() if rec.get("location") else None,
                "event_date": _date(rec.get("event_date")),
                "end_date": _date(rec.get("end_date")),
                "housing_required": bool(housing_bool) if housing_bool is not None else False,
                "booking_link": str(rec["booking_link"]).strip() if rec.get("booking_link") else None,
                "booking_release_at": _datetime(rec.get("booking_release_at")) or _date(rec.get("booking_release_at")),
                "notes": str(rec["notes"]).strip() if rec.get("notes") else None,
            })
    return out


# ---------------------------------------------------------------------------
# Travel  (one row → 1..3 bookings: hotel/car/flight)
# ---------------------------------------------------------------------------
def parse_travel(filename: str, content: bytes) -> List[Dict[str, Any]]:
    sheets = read_table(filename, content)
    out: List[Dict[str, Any]] = []
    for rows in sheets:
        if not rows:
            continue
        hdr_idx = _find_header_row(rows, TRAVEL_HEADERS)
        headers = [str(c) if c is not None else "" for c in rows[hdr_idx]]
        for r in rows[hdr_idx + 1:]:
            rec = _row_to_dict(headers, r, TRAVEL_HEADERS)
            if not rec.get("competition"):
                continue
            bookings = []
            if rec.get("hotel_provider") or rec.get("check_in") or rec.get("hotel_cost"):
                bookings.append({
                    "type": "hotel",
                    "provider": (str(rec["hotel_provider"]).strip() if rec.get("hotel_provider") else None),
                    "confirmation": (str(rec["hotel_confirmation"]).strip() if rec.get("hotel_confirmation") else None),
                    "check_in": _date(rec.get("check_in")),
                    "check_out": _date(rec.get("check_out")),
                    "cancel_by": _date(rec.get("cancel_by")),
                    "cost": _num(rec.get("hotel_cost")) or 0,
                    "amount_paid": _num(rec.get("hotel_paid")) or 0,
                    "balance_due_date": _date(rec.get("hotel_due")),
                })
            if rec.get("car_provider") or rec.get("car_cost") or rec.get("car_pickup_date") or rec.get("car_dropoff_date"):
                pickup_at = _combine_dt(rec.get("car_pickup_date"), rec.get("car_pickup_time"))
                dropoff_at = _combine_dt(rec.get("car_dropoff_date"), rec.get("car_dropoff_time"))
                bookings.append({
                    "type": "car",
                    "provider": (str(rec["car_provider"]).strip() if rec.get("car_provider") else None),
                    "confirmation": (str(rec["car_confirmation"]).strip() if rec.get("car_confirmation") else None),
                    "cost": _num(rec.get("car_cost")) or 0,
                    "amount_paid": 0,
                    "pickup_at": pickup_at,
                    "pickup_location": (str(rec["car_pickup_location"]).strip() if rec.get("car_pickup_location") else None),
                    "dropoff_at": dropoff_at,
                    "dropoff_location": (str(rec["car_dropoff_location"]).strip() if rec.get("car_dropoff_location") else None),
                })
            flight_flag = _bool_from(rec.get("flight_flag"))
            has_flight = (
                rec.get("airline") or rec.get("flight_number") or rec.get("flight_cost")
                or rec.get("outbound_cost") or rec.get("return_cost")
                or rec.get("return_flight_number") or rec.get("return_airline")
                or flight_flag is True
            )
            if has_flight:
                ob = _num(rec.get("outbound_cost"))
                rt = _num(rec.get("return_cost"))
                total = _num(rec.get("flight_cost"))
                if total is None and (ob is not None or rt is not None):
                    total = (ob or 0) + (rt or 0)
                bookings.append({
                    "type": "flight",
                    "provider": (str(rec["airline"]).strip() if rec.get("airline") else None),
                    "confirmation": (str(rec["flight_confirmation"]).strip() if rec.get("flight_confirmation") else None),
                    "flight_number": (str(rec["flight_number"]).strip() if rec.get("flight_number") else None),
                    "depart_time": (str(rec["depart_time"]).strip() if rec.get("depart_time") else None),
                    "arrive_time": (str(rec["arrive_time"]).strip() if rec.get("arrive_time") else None),
                    "outbound_cost": ob,
                    "return_airline": (str(rec["return_airline"]).strip() if rec.get("return_airline") else None),
                    "return_confirmation": (str(rec["return_confirmation"]).strip() if rec.get("return_confirmation") else None),
                    "return_flight_number": (str(rec["return_flight_number"]).strip() if rec.get("return_flight_number") else None),
                    "return_depart_time": (str(rec["return_depart_time"]).strip() if rec.get("return_depart_time") else None),
                    "return_arrive_time": (str(rec["return_arrive_time"]).strip() if rec.get("return_arrive_time") else None),
                    "return_cost": rt,
                    "cost": total or 0,
                    "amount_paid": _num(rec.get("flight_paid")) or 0,
                })
            if not bookings:
                continue
            out.append({
                "competition": str(rec["competition"]).strip(),
                "bookings": bookings,
            })
    return out


# ---------------------------------------------------------------------------
# Schedule (practices, lessons, classes, etc.)
# ---------------------------------------------------------------------------
EVENT_TYPE_MAP = {
    "practice": "practice", "team practice": "practice",
    "team bonding": "team_bonding", "bonding": "team_bonding", "team building": "team_bonding",
    "private lesson": "private_lesson", "private": "private_lesson", "lesson": "private_lesson",
    "choreography": "choreography", "choreo": "choreography",
    "class": "class", "tumbling": "class", "stretching": "class",
}

DOW_MAP = {
    "sun": 0, "sunday": 0,
    "mon": 1, "monday": 1,
    "tue": 2, "tues": 2, "tuesday": 2,
    "wed": 3, "weds": 3, "wednesday": 3,
    "thu": 4, "thur": 4, "thurs": 4, "thursday": 4,
    "fri": 5, "friday": 5,
    "sat": 6, "saturday": 6,
}

FREQ_MAP = {
    "": None, "none": None, "no": None, "never": None,
    "daily": "daily", "every day": "daily",
    "weekly": "weekly", "every week": "weekly",
    "biweekly": "biweekly", "bi-weekly": "biweekly", "bi weekly": "biweekly",
    "every other week": "biweekly", "fortnightly": "biweekly",
    "monthly": "monthly", "every month": "monthly",
}


def _time24(v: Any) -> Optional[str]:
    """Return 'HH:MM' (24h) if parseable, else None.

    Accepts '7:30 PM', '19:30', '7:30pm', '07:30', etc.
    """
    if v is None or v == "":
        return None
    s = str(v).strip()
    if not s:
        return None
    # Try common formats
    for fmt in ("%I:%M %p", "%I:%M%p", "%H:%M", "%I %p", "%I%p"):
        try:
            dt = datetime.strptime(s.upper().replace(" PM", " PM").replace(" AM", " AM"), fmt)
            return dt.strftime("%H:%M")
        except ValueError:
            continue
    # Loose match: digits + optional am/pm
    m = re.match(r"^\s*(\d{1,2})(?::(\d{2}))?\s*(am|pm)?\s*$", s, re.IGNORECASE)
    if m:
        h = int(m.group(1))
        mm = int(m.group(2) or "0")
        ap = (m.group(3) or "").lower()
        if ap == "pm" and h < 12:
            h += 12
        elif ap == "am" and h == 12:
            h = 0
        if 0 <= h <= 23 and 0 <= mm <= 59:
            return f"{h:02d}:{mm:02d}"
    return None


def _parse_days_of_week(v: Any) -> List[int]:
    if v is None or v == "":
        return []
    s = str(v).strip().lower()
    if not s:
        return []
    out: List[int] = []
    for part in re.split(r"[,;/\s]+", s):
        if not part:
            continue
        if part in DOW_MAP:
            out.append(DOW_MAP[part])
    return sorted(set(out))


def parse_schedule(filename: str, content: bytes) -> List[Dict[str, Any]]:
    sheets = read_table(filename, content)
    out: List[Dict[str, Any]] = []
    for rows in sheets:
        if not rows:
            continue
        hdr_idx = _find_header_row(rows, SCHEDULE_HEADERS)
        headers = [str(c) if c is not None else "" for c in rows[hdr_idx]]
        for r in rows[hdr_idx + 1:]:
            rec = _row_to_dict(headers, r, SCHEDULE_HEADERS)
            title = (str(rec["title"]).strip() if rec.get("title") else "")
            event_date = _date(rec.get("date"))
            if not title or not event_date:
                continue

            etype_raw = (str(rec.get("event_type") or "")).strip().lower()
            event_type = EVENT_TYPE_MAP.get(etype_raw, etype_raw if etype_raw in (
                "practice", "team_bonding", "private_lesson", "choreography", "class", "other"
            ) else "practice")

            athletes_raw = rec.get("athletes")
            athlete_names: List[str] = []
            if athletes_raw:
                for n in re.split(r"[,;|]+", str(athletes_raw)):
                    nn = n.strip()
                    if nn:
                        athlete_names.append(nn)

            freq = FREQ_MAP.get((str(rec.get("repeats") or "").strip().lower()), None)
            recurrence_rule: Optional[Dict[str, Any]] = None
            if freq:
                until = _date(rec.get("repeat_until"))
                if until:
                    recurrence_rule = {
                        "frequency": freq,
                        "days_of_week": _parse_days_of_week(rec.get("repeat_days")),
                        "until": until,
                    }

            out.append({
                "title": title,
                "event_type": event_type,
                "date": event_date,
                "start_time": _time24(rec.get("start_time")),
                "end_time": _time24(rec.get("end_time")),
                "location": (str(rec["location"]).strip() if rec.get("location") else None),
                "athlete_names": athlete_names,
                "notes": (str(rec["notes"]).strip() if rec.get("notes") else None),
                "recurrence_rule": recurrence_rule,
            })
    return out


# ---------------------------------------------------------------------------
# Expenses (supports long-form and wide-form)
# ---------------------------------------------------------------------------
def parse_expenses(filename: str, content: bytes) -> Dict[str, Any]:
    """Returns {format: 'long'|'wide', rows: [...], athlete_columns: [...]}

    - Long form rows: {date, athlete, category, amount, due_date, paid, note}
    - Wide form rows: {date, category, amounts: {<col>: amount}}  (categories = rows; athletes = columns)
    - Athlete-grid form: detected when columns are categories AND first column is Month.
      Emits long-form rows with auto-detected athlete from banner / sheet name.
    """
    sheets = read_table(filename, content)
    long_rows: List[Dict[str, Any]] = []
    wide_rows: List[Dict[str, Any]] = []
    wide_columns: List[str] = []

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet_names = wb.sheetnames
    except Exception:
        sheet_names = []

    for sheet_idx, rows in enumerate(sheets):
        if not rows:
            continue
        hdr_idx = _find_header_row(rows, EXPENSE_HEADERS)
        headers_raw = [str(c) if c is not None else "" for c in rows[hdr_idx]]
        norm_headers = [_norm(h) for h in headers_raw]
        known_count = sum(1 for h in norm_headers if h in EXPENSE_HEADERS)
        canonical = [EXPENSE_HEADERS.get(h) for h in norm_headers]
        is_long = known_count >= 3 and "athlete" in canonical

        # Athlete-grid detection: first column resolves to date/month AND >=3 categories follow
        first_canonical = canonical[0] if canonical else None
        category_cols = [(i, headers_raw[i].strip()) for i, h in enumerate(headers_raw)
                         if i > 0 and h.strip() and h.strip() not in ("TOTAL", "Total")]
        looks_like_grid = (
            first_canonical == "date"
            and len(category_cols) >= 3
            and not is_long
        )

        if is_long:
            for r in rows[hdr_idx + 1:]:
                rec = _row_to_dict(headers_raw, r, EXPENSE_HEADERS)
                if not (rec.get("athlete") and rec.get("category") and _num(rec.get("amount"))):
                    continue
                long_rows.append({
                    "date": _date(rec.get("date")),
                    "athlete": str(rec["athlete"]).strip(),
                    "category": str(rec["category"]).strip(),
                    "amount": _num(rec.get("amount")) or 0,
                    "due_date": _date(rec.get("due_date")),
                    "paid": bool(_bool_from(rec.get("paid"))) if rec.get("paid") is not None else False,
                    "note": str(rec["note"]).strip() if rec.get("note") else None,
                })
            continue

        if looks_like_grid:
            # Detect athlete from banner (first non-empty cell above headers) or sheet name
            athlete_name: Optional[str] = None
            for prev in rows[:hdr_idx]:
                for c in prev:
                    if c and str(c).strip():
                        candidate = str(c).strip()
                        # strip generic "NAME" suffix
                        cleaned = re.sub(r"\s*name\s*$", "", candidate, flags=re.I).strip()
                        if cleaned:
                            athlete_name = cleaned
                            break
                if athlete_name:
                    break
            if not athlete_name and 0 <= sheet_idx < len(sheet_names):
                sn = sheet_names[sheet_idx]
                if sn and sn.lower() not in ("sheet1", "sheet2"):
                    athlete_name = sn
            if not athlete_name:
                athlete_name = f"Athlete #{sheet_idx + 1}"

            for r in rows[hdr_idx + 1:]:
                if not r or r[0] in (None, ""):
                    continue
                row_date = _date(r[0]) or _parse_month_year(str(r[0]))
                for col_idx, col_name in category_cols:
                    if col_idx >= len(r):
                        continue
                    amt = _num(r[col_idx])
                    if amt is None or amt <= 0:
                        continue
                    long_rows.append({
                        "date": row_date,
                        "athlete": athlete_name,
                        "category": col_name,
                        "amount": amt,
                        "due_date": None,
                        "paid": False,
                        "note": None,
                    })
            continue

        # Wide form: categories down rows, athletes across columns
        category_idx = None
        for i, h in enumerate(norm_headers):
            if EXPENSE_HEADERS.get(h) == "category" or h in ("", "expenses", "expense", "type"):
                category_idx = i
                break
        if category_idx is None:
            category_idx = 0

        athlete_cols_idx: List[int] = []
        for i, h in enumerate(headers_raw):
            if i == category_idx:
                continue
            if EXPENSE_HEADERS.get(_norm(h)):
                continue
            if not h or not h.strip():
                continue
            if h.strip().lower() in ("total", "totals", "subtotal"):
                continue
            athlete_cols_idx.append(i)

        if not athlete_cols_idx:
            continue
        for i in athlete_cols_idx:
            col_name = headers_raw[i].strip()
            if col_name and col_name not in wide_columns:
                wide_columns.append(col_name)

        sheet_date: Optional[str] = None
        if 0 <= sheet_idx < len(sheet_names):
            sheet_date = _parse_month_year(sheet_names[sheet_idx])

        for r in rows[hdr_idx + 1:]:
            if category_idx >= len(r):
                continue
            cat_val = r[category_idx]
            if not cat_val or not str(cat_val).strip():
                continue
            category = str(cat_val).strip()
            if category.lower() in ("total", "totals", "subtotal", "balance"):
                continue
            amounts: Dict[str, float] = {}
            for i in athlete_cols_idx:
                if i >= len(r):
                    continue
                n = _num(r[i])
                if n is not None and n > 0:
                    amounts[headers_raw[i].strip()] = n
            if not amounts:
                continue
            wide_rows.append({
                "date": sheet_date,
                "category": category,
                "amounts": amounts,
            })

    if long_rows:
        return {"format": "long", "rows": long_rows, "athlete_columns": []}
    return {"format": "wide", "rows": wide_rows, "athlete_columns": wide_columns}


def _parse_month_year(s: str) -> Optional[str]:
    """Parse strings like 'August 2025', 'Aug 25', '2025-08', 'Aug' (assume current year)."""
    s = s.strip()
    for fmt in ("%B %Y", "%b %Y", "%B %y", "%b %y", "%Y-%m", "%m/%Y"):
        try:
            d = datetime.strptime(s, fmt).date().replace(day=1)
            return d.isoformat()
        except ValueError:
            continue
    # Try just month name with current year
    for fmt in ("%B", "%b"):
        try:
            month = datetime.strptime(s, fmt).month
            return date(datetime.now().year, month, 1).isoformat()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Teams to Watch (one row → one team to watch, matched to a competition)
# ---------------------------------------------------------------------------
def parse_teams_to_watch(filename: str, content: bytes) -> List[Dict[str, Any]]:
    sheets = read_table(filename, content)
    out: List[Dict[str, Any]] = []
    for rows in sheets:
        if not rows:
            continue
        hdr_idx = _find_header_row(rows, TEAMS_TO_WATCH_HEADERS)
        headers = [str(c) if c is not None else "" for c in rows[hdr_idx]]
        for r in rows[hdr_idx + 1:]:
            rec = _row_to_dict(headers, r, TEAMS_TO_WATCH_HEADERS)
            name = (str(rec["name"]).strip() if rec.get("name") else "")
            if not name:
                continue
            out.append({
                "competition": (str(rec["competition"]).strip() if rec.get("competition") else None),
                "name": name,
                "date": _date(rec.get("date")),
                "location": (str(rec["location"]).strip() if rec.get("location") else None),
                "performance_time": _time24(rec.get("performance_time")),
            })
    return out


# ---------------------------------------------------------------------------
# Team Hub — Roster / Sizes / Paperwork / Payments
# ---------------------------------------------------------------------------
ROSTER_HEADERS = {
    "first name": "first_name", "first": "first_name",
    "last name": "last_name", "last": "last_name",
    "name": "name", "full name": "name", "member": "name", "athlete": "name", "athlete name": "name",
    "role": "role", "type": "role", "position": "role",
    "phone": "phone", "cell": "phone", "mobile": "phone", "phone number": "phone", "cell phone": "phone",
    "email": "email", "email address": "email", "e mail": "email",
    "parent first name": "parent_first_name", "parent first": "parent_first_name", "guardian first name": "parent_first_name",
    "parent last name": "parent_last_name", "parent last": "parent_last_name", "guardian last name": "parent_last_name",
    "parent name": "parent_name", "guardian name": "parent_name", "parent guardian": "parent_name",
    "parent phone": "parent_phone", "guardian phone": "parent_phone", "parent cell": "parent_phone", "parent mobile": "parent_phone",
    "parent email": "parent_email", "guardian email": "parent_email",
    "team": "teams", "teams": "teams", "team s": "teams", "team name": "teams", "team names": "teams",
    "notes": "notes", "note": "notes",
}

TEAM_PAYMENT_HEADERS = {
    "name": "name", "member": "name", "athlete": "name", "full name": "name",
    "amount paid": "amount_paid", "amount": "amount_paid", "paid amount": "amount_paid", "amt": "amount_paid",
    "method": "method", "payment method": "method", "pay method": "method", "how": "method",
    "date paid": "paid_on", "paid date": "paid_on", "date": "paid_on", "paid on": "paid_on",
    "paid": "paid", "status": "paid", "paid status": "paid",
}

_ROLE_MAP = {
    "athlete": "athlete", "cheerleader": "athlete", "cheer": "athlete", "kid": "athlete",
    "coach": "coach", "head coach": "coach", "asst coach": "coach", "assistant coach": "coach",
    "team rep": "team_rep", "rep": "team_rep", "manager": "team_rep", "team manager": "team_rep", "team parent": "team_rep",
    "staff": "staff", "admin": "staff", "choreographer": "staff",
    "parent": "parent", "guardian": "parent", "mom": "parent", "dad": "parent",
}


def _norm_role(v: Any) -> str:
    return _ROLE_MAP.get(_norm(v), "athlete")


def parse_roster(filename: str, content: bytes) -> List[Dict[str, Any]]:
    sheets = read_table(filename, content)
    out: List[Dict[str, Any]] = []
    for rows in sheets:
        if not rows:
            continue
        hdr_idx = _find_header_row(rows, ROSTER_HEADERS)
        headers = [str(c) if c is not None else "" for c in rows[hdr_idx]]
        for r in rows[hdr_idx + 1:]:
            rec = _row_to_dict(headers, r, ROSTER_HEADERS)
            first = str(rec.get("first_name") or "").strip()
            last = str(rec.get("last_name") or "").strip()
            name = str(rec.get("name") or "").strip() or f"{first} {last}".strip()
            if not name:
                continue
            if not first and not last:
                parts = name.split()
                first = parts[0]
                last = " ".join(parts[1:]) if len(parts) > 1 else ""
            # split combined parent name if provided
            pfn = str(rec.get("parent_first_name") or "").strip()
            pln = str(rec.get("parent_last_name") or "").strip()
            if not pfn and not pln and rec.get("parent_name"):
                pp = str(rec["parent_name"]).strip().split()
                if pp:
                    pfn = pp[0]
                    pln = " ".join(pp[1:]) if len(pp) > 1 else ""
            teams_raw = str(rec.get("teams") or "").strip()
            team_names = [t.strip() for t in re.split(r"[;,/|]", teams_raw) if t.strip()] if teams_raw else []
            out.append({
                "name": name,
                "first_name": first or None,
                "last_name": last or None,
                "role": _norm_role(rec.get("role")),
                "phone": str(rec["phone"]).strip() if rec.get("phone") else None,
                "email": str(rec["email"]).strip() if rec.get("email") else None,
                "parent_first_name": pfn or None,
                "parent_last_name": pln or None,
                "parent_phone": str(rec["parent_phone"]).strip() if rec.get("parent_phone") else None,
                "parent_email": str(rec["parent_email"]).strip() if rec.get("parent_email") else None,
                "team_names": team_names,
                "notes": str(rec["notes"]).strip() if rec.get("notes") else None,
            })
    return out


_GRID_NAME_KEYS = {"name", "member", "athlete", "player", "full name", "athlete name"}


def parse_named_grid(filename: str, content: bytes) -> Dict[str, Any]:
    """Wide sheet: one column is the person's name, the rest become data columns.

    Returns {"columns": [labels], "rows": [{"name": str, "cells": {label: value}}]}.
    Used by Sizes and Paperwork imports.
    """
    sheets = read_table(filename, content)
    for rows in sheets:
        if not rows:
            continue
        # Header = first row that has >=2 non-empty cells.
        hdr_idx = 0
        for i, row in enumerate(rows[:10]):
            if sum(1 for c in row if c not in (None, "")) >= 2:
                hdr_idx = i
                break
        headers = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
        # Locate the name column.
        name_idx = None
        first_idx = last_idx = None
        for i, h in enumerate(headers):
            n = _norm(h)
            if n in _GRID_NAME_KEYS and name_idx is None:
                name_idx = i
            if n in ("first name", "first"):
                first_idx = i
            if n in ("last name", "last"):
                last_idx = i
        if name_idx is None and first_idx is None:
            name_idx = 0
        skip = {i for i in (name_idx, first_idx, last_idx) if i is not None}
        columns = [headers[i] for i in range(len(headers)) if i not in skip and headers[i]]
        out_rows: List[Dict[str, Any]] = []
        for r in rows[hdr_idx + 1:]:
            vals = list(r) + [None] * (len(headers) - len(r))
            if name_idx is not None:
                name = str(vals[name_idx]).strip() if vals[name_idx] is not None else ""
            else:
                fn = str(vals[first_idx]).strip() if first_idx is not None and vals[first_idx] is not None else ""
                ln = str(vals[last_idx]).strip() if last_idx is not None and vals[last_idx] is not None else ""
                name = f"{fn} {ln}".strip()
            if not name:
                continue
            cells: Dict[str, Any] = {}
            for i, h in enumerate(headers):
                if i in skip or not h:
                    continue
                v = vals[i]
                if v not in (None, ""):
                    cells[h] = str(v).strip()
            out_rows.append({"name": name, "cells": cells})
        return {"columns": columns, "rows": out_rows}
    return {"columns": [], "rows": []}


def parse_team_payments(filename: str, content: bytes) -> List[Dict[str, Any]]:
    sheets = read_table(filename, content)
    out: List[Dict[str, Any]] = []
    for rows in sheets:
        if not rows:
            continue
        hdr_idx = _find_header_row(rows, TEAM_PAYMENT_HEADERS)
        headers = [str(c) if c is not None else "" for c in rows[hdr_idx]]
        for r in rows[hdr_idx + 1:]:
            rec = _row_to_dict(headers, r, TEAM_PAYMENT_HEADERS)
            name = str(rec.get("name") or "").strip()
            if not name:
                continue
            amt = _num(rec.get("amount_paid"))
            paid_flag = _bool_from(rec.get("paid"))
            paid = bool(paid_flag) if paid_flag is not None else (amt is not None and amt > 0)
            out.append({
                "name": name,
                "amount_paid": amt,
                "method": str(rec["method"]).strip() if rec.get("method") else None,
                "paid_on": _date(rec.get("paid_on")),
                "paid": paid,
            })
    return out


# ---------------------------------------------------------------------------
# CSV Templates
# ---------------------------------------------------------------------------
TEMPLATES: Dict[str, Tuple[List[str], List[List[str]]]] = {
    "competitions": (
        ["Competition Name", "Location", "Event Date", "End Date",
         "Housing Required", "Booking Link", "Booking Release Date and Time", "Notes"],
        [
            ["NCA Senior Nationals", "Houston, TX", "2025-11-13", "2025-11-15",
             "Yes", "https://stayandplay.example/nca", "2025-09-01 10:00", "Pack uniform Friday"],
            ["Spirit Sports Palm Springs", "Palm Springs, CA", "2025-08-26", "2025-08-27",
             "No", "", "", ""],
        ],
    ),
    "travel": (
        ["Competition", "Hotel Name", "Hotel Confirmation", "Check In", "Check Out",
         "Cancel Date", "Hotel Cost", "Hotel Paid", "Balance Due Date",
         "Rental Car Company", "Rental Car Confirmation", "Rental Car Cost",
         "Pickup Date", "Pickup Time", "Pickup Location",
         "Dropoff Date", "Dropoff Time", "Dropoff Location",
         "Airline", "Flight Confirmation", "Flight Number",
         "Depart Time", "Arrive Time", "Outbound Cost",
         "Return Airline", "Return Confirmation", "Return Flight Number",
         "Return Depart Time", "Return Arrive Time", "Return Cost",
         "Flight Cost", "Flight Paid"],
        [
            ["NCA Senior Nationals", "Hyatt Regency Houston", "ABC123", "2025-11-13", "2025-11-15",
             "2025-10-29", "650.00", "200.00", "2025-10-29",
             "Enterprise", "RES7788", "180.00",
             "2025-11-13", "12:30 PM", "Houston Airport",
             "2025-11-16", "3:00 PM", "Houston Airport",
             "Southwest", "WN42X", "WN1234",
             "2025-11-13 08:30", "2025-11-13 10:15", "160.00",
             "Southwest", "WN42Y", "WN5678",
             "2025-11-16 16:00", "2025-11-16 18:30", "160.00",
             "320.00", "320.00"],
        ],
    ),
    "expenses": (
        ["Date", "Athlete", "Category", "Amount", "Due Date", "Paid", "Note"],
        [
            ["2025-10-01", "Ava", "Tuition", "250.00", "2025-10-05", "No", "October tuition"],
            ["2025-10-01", "Ava", "Comp/Choreo", "75.00", "", "Yes", ""],
            ["2025-10-01", "Mia", "Tuition", "250.00", "2025-10-05", "No", ""],
        ],
    ),
    "schedule": (
        ["Title", "Type", "Date", "Start Time", "End Time", "Location",
         "Athletes", "Repeats", "Repeat Days", "Repeat Until", "Notes"],
        [
            ["Senior 5 practice", "Practice", "2025-09-02", "6:30 PM", "8:30 PM",
             "California Allstars", "Ava, Mia", "Weekly", "Tue,Thu", "2025-12-16",
             "Wear comp shoes"],
            ["Private tumbling", "Private Lesson", "2025-09-05", "4:00 PM", "5:00 PM",
             "Gym B", "Ava", "Weekly", "Fri", "2025-11-28", ""],
            ["Team bonding pizza", "Team Bonding", "2025-09-13", "7:00 PM", "9:00 PM",
             "Coach's house", "Ava, Mia", "", "", "", "Bring drink"],
        ],
    ),
    "teams_to_watch": (
        ["Competition", "Team Name", "Date", "Location", "Performance Time"],
        [
            ["NCA Senior Nationals", "Cheer Athletics Panthers", "2025-11-14", "Arena Floor A", "2:30 PM"],
            ["NCA Senior Nationals", "Top Gun Large Coed", "2025-11-15", "Arena Floor B", "11:00 AM"],
            ["Spirit Sports Palm Springs", "California Allstars Smoed", "2025-08-26", "Main Stage", "4:15 PM"],
        ],
    ),
    "roster": (
        ["First Name", "Last Name", "Role", "Phone", "Email",
         "Parent First Name", "Parent Last Name", "Parent Phone", "Parent Email", "Team(s)", "Notes"],
        [
            ["Ava", "Johnson", "Athlete", "", "", "Sarah", "Johnson", "(555) 123-4567", "sarah@example.com", "Senior 5", "Flyer"],
            ["Mia", "Lopez", "Athlete", "", "", "Carlos", "Lopez", "(555) 222-3344", "carlos@example.com", "Senior 5", ""],
            ["Coach Kim", "Reed", "Coach", "(555) 987-6543", "kim@gym.com", "", "", "", "", "Senior 5, Youth 1", ""],
        ],
    ),
    "team_sizes": (
        ["Name", "Shirt", "Shorts", "Shoes", "Jacket"],
        [
            ["Ava Johnson", "YL", "YM", "3", "AS"],
            ["Mia Lopez", "AS", "AS", "5", "AS"],
            ["Coach Kim", "AL", "AM", "8", "AM"],
        ],
    ),
    "team_paperwork": (
        ["Name", "Physical Form", "Waiver", "Media Release"],
        [
            ["Ava Johnson", "Yes", "Yes", "No"],
            ["Mia Lopez", "Yes", "No", "No"],
            ["Coach Kim", "Yes", "Yes", "Yes"],
        ],
    ),
    "team_payments": (
        ["Name", "Amount Paid", "Method", "Date Paid", "Paid"],
        [
            ["Ava Johnson", "150.00", "Venmo", "2025-10-05", "Yes"],
            ["Mia Lopez", "75.00", "Cash", "2025-10-06", "Yes"],
            ["Coach Kim", "", "", "", "No"],
        ],
    ),
}


# Valid option reference shown on the XLSX "Reference" sheet for each kind.
TEMPLATE_NOTES: Dict[str, List[str]] = {
    "expenses": [
        "Valid Category values:",
        *[f"  • {c}" for c in KNOWN_CATEGORIES],
        "",
        "Paid column accepts: Yes / No (also true/false, paid/unpaid, 1/0).",
        "Date / Due Date format: YYYY-MM-DD (e.g. 2025-10-05).",
    ],
    "competitions": [
        "Housing Required accepts: Yes / No.",
        "Dates use YYYY-MM-DD. Booking Release accepts 'YYYY-MM-DD HH:MM'.",
    ],
    "schedule": [
        "Type values: Practice, Team Bonding, Private Lesson, Choreography, Class, Other.",
        "Repeats values: Weekly, Biweekly, Monthly, Daily (leave blank for one-off).",
        "Repeat Days: comma-separated day names, e.g. 'Tue,Thu'.",
        "Times accept 12h (6:30 PM) or 24h (18:30).",
    ],
    "travel": [
        "One row per competition. Fill only the columns you have.",
        "Dates use YYYY-MM-DD; times accept 12h or 24h.",
    ],
    "teams_to_watch": [
        "Competition must match (or will create) a competition by name.",
        "Date uses YYYY-MM-DD. Performance Time accepts 12h (2:30 PM) or 24h.",
    ],
    "roster": [
        "Role values: Athlete, Coach, Team Rep (or Manager), Staff, Parent. Blank = Athlete.",
        "Team(s): separate multiple teams with commas, e.g. 'Senior 5, Youth 1'. New team names are created automatically.",
        "Re-importing a person with the same name updates their existing roster entry.",
    ],
    "team_sizes": [
        "First column is the person's Name; every other column becomes a size category.",
        "Add any columns you like (Shirt, Shorts, Shoes, Bow, Ring…). Values are free text.",
        "Names are matched to your roster; unmatched names are added as new Athletes.",
    ],
    "team_paperwork": [
        "First column is the person's Name; every other column becomes a check-off item.",
        "Cell values accept Yes / No (also true/false, x, done, 1/0).",
        "Each import creates a new paperwork sheet. Unmatched names are added as new Athletes.",
    ],
    "team_payments": [
        "Paid accepts Yes / No. Amount Paid is a number; Date Paid uses YYYY-MM-DD.",
        "Each import creates a new payment tracker. Unmatched names are added as new Athletes.",
    ],
}


def render_template_csv(kind: str) -> str:
    if kind not in TEMPLATES:
        raise ValueError(f"Unknown template: {kind}")
    headers, rows = TEMPLATES[kind]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    return buf.getvalue()


def render_template_xlsx(kind: str) -> bytes:
    if kind not in TEMPLATES:
        raise ValueError(f"Unknown template: {kind}")
    from openpyxl import Workbook
    from openpyxl.styles import Font

    headers, rows = TEMPLATES[kind]
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append(list(r))
    # Reasonable default column widths
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(str(h)) + 2)

    notes = TEMPLATE_NOTES.get(kind)
    if notes:
        ref = wb.create_sheet("Reference")
        ref.append(["How to fill this template"])
        ref["A1"].font = Font(bold=True)
        for line in notes:
            ref.append([line])
        ref.column_dimensions["A"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
